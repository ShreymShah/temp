from flask import Flask, render_template, request, redirect, url_for, flash
import time as t
import threading
from pya3 import *
import openpyxl

app = Flask(__name__,template_folder='template')
app.secret_key = "DONT"

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/add_account", methods=["POST", "GET"])
def add_account():
    if request.method == "POST":
        username = request.form["username"]
        api_key = request.form["api_key"]
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        sheet.cell(row=sheet.max_row+1,column=1).value = username
        sheet.cell(row=sheet.max_row, column=2).value = api_key
        wb.save("data.xlsx")
        flash("account added")
        return redirect(url_for("add_account"))

    else:
        return render_template("add_account.html")

@app.route("/accounts")
def accounts():
    wb = openpyxl.load_workbook("data.xlsx")
    sheet = wb.active
    usernames = []
    for i in range(2,sheet.max_row+1):
        cell = sheet.cell(row=i, column=1).value
        if cell!=None:
            usernames.append(cell)

    return render_template("accounts.html",username=usernames)

@app.route("/delete_account/<username>")
def delete_account(username):
    print(username)
    wb = openpyxl.load_workbook("data.xlsx")
    sheet = wb.active
    for i in range(1,sheet.max_row+1):
        print(sheet.cell(row=i,column=1).value)
        if username == sheet.cell(row=i,column=1).value:
            sheet.cell(row=i, column=1).value = ""
            sheet.cell(row=i, column=2).value = ""
            wb.save("data.xlsx")
            flash("Account Deleted!")
    return redirect(url_for("accounts"))

@app.route("/new_trade",methods=["POST","GET"])
def new_trade():
    if request.method == "POST":
        dic_qty = {}
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = {}
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1).value
            if cell != None:
                usernames[cell] = sheet.cell(row=i, column=2).value
        call_sell = request.form["call_sell"]
        call_buy = request.form["call_buy"]
        put_sell = request.form["put_sell"]
        put_buy = request.form["put_buy"]
        expiry = request.form["expiry"]
        
        for user in usernames:
            dic_qty[user] = request.form[user]
        try:
            for key,value in dic_qty.items():
                if value!='0':
                    t1 = threading.Thread(target=take_new_trade, args=(key,usernames[key],call_sell,call_buy,put_sell,put_buy,value,expiry,))
                    t1.start()
            flash("Trades Taken!")
        except Exception as e:
            print(e)
            flash("Some error occurred!")
        return redirect(url_for("new_trade"))

    else:
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = []
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1).value
            if cell != None:
                usernames.append(cell)
        return render_template("new_trade.html",usernames=usernames)

@app.route("/shifting",methods=["POST","GET"])
def shifting():
    if request.method == "POST":
        dic_qty = {}
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = {}
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1).value
            if cell != None:
                usernames[cell] = sheet.cell(row=i, column=2).value

        previous_call_sold = request.form["current_call_sold"]
        previous_call_hedge = request.form["current_call_hedge"]
        new_call_sell = request.form["new_call_sell"]
        new_call_hedge = request.form["new_call_hedge"]

        previous_put_sold = request.form["current_put_sold"]
        previous_put_hedge = request.form["current_put_hedge"]
        new_put_sell = request.form["new_put_sell"]
        new_put_hedge = request.form["new_put_hedge"]

        previous_expiry = request.form["previous_expiry"]
        current_expiry = request.form["current_expiry"]

        for key,value in usernames.items():
            dic_qty[key] = request.form[key]

        for key,value in dic_qty.items():
            if value!='0':
                t1 = threading.Thread(target=shift, args=(key,usernames[key],previous_call_sold, new_call_sell,previous_call_hedge,new_call_hedge, previous_put_sold, new_put_sell,previous_put_hedge,new_put_hedge,previous_expiry,current_expiry,value,))
                t1.start()
        flash("Shifting Done!")
        return redirect(url_for("shifting"))

    else:
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = []
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1).value
            if cell != None:
                usernames.append(cell)
        return render_template("shifting.html",usernames=usernames)

def take_new_trade(username, api_key, call_sell, call_buy, put_sell, put_buy, qty,expiry):
    qty = int(qty)
    alice = Aliceblue(user_id=username, api_key=api_key)
    aliceblue_Res = alice.get_session_id()
    print(aliceblue_Res)
    alice.get_contract_master("NFO")
    a = int(qty / 1800)
    for i in range(0, a):
        PlaceBuyOrder(alice,1800,True,call_buy,expiry)
        PlaceSellOrder(alice, 1800, True, call_sell, expiry)
        PlaceBuyOrder(alice,1800,False,put_buy,expiry)
        PlaceSellOrder(alice, 1800, False, put_sell, expiry)

    PlaceBuyOrder(alice, qty-(1800*a), True, call_buy, expiry)
    PlaceSellOrder(alice, qty-(1800*a), True, call_sell, expiry)
    PlaceBuyOrder(alice, qty-(1800*a), False, put_buy, expiry)
    PlaceSellOrder(alice, qty-(1800*a), False, put_sell, expiry)

def shift(username, api_key, previous_call_sold, new_call_sell,previous_call_hedge,new_call_hedge, previous_put_sold, new_put_sell,previous_put_hedge,new_put_hedge,previous_expiry,current_expiry,qty):
    qty = int(qty)
    alice = Aliceblue(user_id=username, api_key=api_key)
    aliceblue_Res = alice.get_session_id()
    print(aliceblue_Res)
    alice.get_contract_master("NFO")

    #squareoff currently sold put
    if previous_put_sold!=new_put_sell:
        a = int(qty / 1800)
        for i in range(0,a):
            PlaceBuyOrder(alice,1800,False,previous_put_sold,previous_expiry)
        PlaceBuyOrder(alice, qty-(1800*a),False,previous_put_sold,previous_expiry)

    #squareoff currently sold call
    if previous_call_sold!=new_call_sell:
        a = int(qty / 1800)
        for i in range(0,a):
            PlaceBuyOrder(alice,1800,True,previous_call_sold,previous_expiry)
        PlaceBuyOrder(alice, qty-(1800*a),True,previous_call_sold,previous_expiry)

    #take new hedge for put
    if previous_put_hedge != new_put_hedge:
        a = int(qty / 1800)
        for i in range(0, a):
            PlaceBuyOrder(alice, 1800, False, new_put_hedge, current_expiry)
            PlaceSellOrder(alice, 1800, False, previous_put_hedge, previous_expiry)
        PlaceBuyOrder(alice, qty - (1800 * a), False, new_put_hedge, current_expiry)
        PlaceSellOrder(alice, qty - (1800 * a), False, previous_put_hedge, previous_expiry)

    #take new hedge for call
    if previous_call_hedge != new_call_hedge:
        a = int(qty / 1800)
        for i in range(0, a):
            PlaceBuyOrder(alice, 1800, True, new_call_hedge, current_expiry)
            PlaceSellOrder(alice, 1800, True, previous_call_hedge, previous_expiry)
        PlaceBuyOrder(alice, qty - (1800 * a), True, new_call_hedge, current_expiry)
        PlaceSellOrder(alice, qty - (1800 * a), True, previous_call_hedge, previous_expiry)

    #sell new put
    if previous_put_sold!=new_put_sell:
        a = int(qty / 1800)
        for i in range(0, a):
            PlaceSellOrder(alice, 1800, False,new_put_sell,current_expiry)
        PlaceSellOrder(alice, qty - (1800 * a), False, new_put_sell, current_expiry)

    #sell new call
    if previous_call_sold!=new_call_sell:
        a = int(qty / 1800)
        for i in range(0, a):
            PlaceSellOrder(alice, 1800, True, new_call_sell, current_expiry)
        PlaceSellOrder(alice, qty - (1800 * a), True, new_call_sell, current_expiry)

def PlaceBuyOrder(alice, qty, call,strike,expiry):
    res_2 = alice.place_order(transaction_type=TransactionType.Buy,
                            instrument=alice.get_instrument_for_fno(exch="NFO",symbol='NIFTY', expiry_date=expiry, is_fut=False,strike=int(strike), is_CE=call),
                            quantity=qty,
                            order_type=OrderType.Market,
                            product_type=ProductType.Normal,
                            price=0.0,
                            trigger_price=None,
                            stop_loss=None,
                            square_off=None,
                            trailing_sl=None,
                            is_amo=False,
                            order_tag='order1')
    t.sleep(2)

def PlaceSellOrder(alice, qty, call,strike,expiry):
    res_2 = alice.place_order(transaction_type=TransactionType.Sell,
                            instrument=alice.get_instrument_for_fno(exch="NFO",symbol='NIFTY', expiry_date=expiry, is_fut=False,strike=int(strike), is_CE=call),
                            quantity=qty,
                            order_type=OrderType.Market,
                            product_type=ProductType.Normal,
                            price=0.0,
                            trigger_price=None,
                            stop_loss=None,
                            square_off=None,
                            trailing_sl=None,
                            is_amo=False,
                            order_tag='order1')
    t.sleep(2)

if __name__ == "__main__":
    app.run(debug=True)
